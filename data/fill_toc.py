import os.path
import shutil
import openpyxl
import config
import time
import json

rootPath = config.rootPath
wxlistfile = rootPath + '\\wxlist.xlsx'
wxlistfilefinal = rootPath + '\\wxlist-final.xlsx'
# wxlist-final-ad
wxlistfilefinal = rootPath + '\\wxlist-final-ad.xlsx'

class Article:
    def __init__(self, data):
        self.aid = data.get("aid")
        self.album_id = data.get("album_id")
        self.appmsg_album_infos = data.get("appmsg_album_infos")
        self.appmsgid = data.get("appmsgid")
        self.checking = data.get("checking")
        self.copyright_type = data.get("copyright_type")
        self.cover = data.get("cover")
        self.create_time = data.get("create_time")
        self.digest = data.get("digest")
        self.has_red_packet_cover = data.get("has_red_packet_cover")
        self.is_pay_subscribe = data.get("is_pay_subscribe")
        self.item_show_type = data.get("item_show_type")
        self.itemidx = data.get("itemidx")
        self.link = data.get("link")
        self.media_duration = data.get("media_duration")
        self.mediaapi_publish_status = data.get("mediaapi_publish_status")
        self.pay_album_info = data.get("pay_album_info")
        self.tagid = data.get("tagid")
        self.title = data.get("title")
        self.update_time = data.get("update_time")

    @property
    def month(self):
        return time.strftime("%Y-%m",time.localtime(int(self.create_time)))
        #return datetime.fromtimestamp(self.update_time).strftime('%Y-%m')

def fill_toc():
    wb = openpyxl.load_workbook(wxlistfile)
    ws = wb.active
    # 读取每一行的值
    data_list = []
    for row in ws.iter_rows(values_only=True):
        data_list.append(row[2])
    data_list.reverse()

    articles = []
    for row in data_list:
        # 解析 JSON 字符串
        data_dict = json.loads(row)
        copyright_type = data_dict.get("copyright_type")
        #if copyright_type == 0:
        #    print(f'广告：{data_dict.get("title")}')
        #    continue
        # 创建对象
        article = Article(data_dict)
        articles.append(article)

    # 按月份分类
    articles_by_month = {}
    for article in articles:
        month = article.month
        if month not in articles_by_month:
            articles_by_month[month] = []
        articles_by_month[month].append(article)

    copyPath = 'D:\\WorkSpace\\idea\\python\\wechat\\toc'
    htmlPath = 'D:\\WorkSpace\\idea\\python\\wechat\\html\\'
    # 打印按月份分类的结果
    for month, articles in articles_by_month.items():
        print(f"# {month}")
        # 复制文件并重命名
        #shutil.copy(copyPath+"\\1.html", copyPath+"\\"+month+".html")
        for article in articles:
            if os.path.exists(htmlPath+article.title+".html"):
                print(f"## {article.title}")

def save_final():
    wb = openpyxl.load_workbook(wxlistfile)
    ws = wb.active
    # 读取每一行的值
    data_list = []
    for row in ws.iter_rows(values_only=True):
        data_list.append(row[2])
    data_list.reverse()

    articles = []
    wb_save = openpyxl.load_workbook(wxlistfilefinal)
    ws_save = wb_save.active

    for row in data_list:
        # 解析 JSON 字符串
        data_dict = json.loads(row)
        copyright_type = data_dict.get("copyright_type")
        if copyright_type == 0:
            create_time = data_dict.get("create_time")
            title = data_dict.get("title")
            link = data_dict.get("link")
            aid = data_dict.get("aid")
            row = [aid, title, time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(int(create_time))), link]
            ws_save.append(row)

    wb_save.save(wxlistfilefinal)

if __name__ == '__main__':
    fill_toc()
    #save_final()


