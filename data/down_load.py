import os
import uuid
import requests
import re
from bs4 import BeautifulSoup
import config
import markdownify
import time
from typing import List, Optional, Dict
import unicodedata
import hashlib
from datetime import datetime
import pandas as pd
import openpyxl
import random

# 给定的 URL
url = "https://mp.weixin.qq.com/s?__biz=MzU1MTk2NDE4Mg==&mid=2247489209&idx=1&sn=67694d6afbb51b0b2a4866cd310224f8&chksm=fb880dc0ccff84d6e16a1a33a726c682659bfd2c6d5e1128440d95dadd9950b201b6db82bb54#rd"


def hash_byte_data(byte_data: bytes) -> str:
    return hashlib.sha256(byte_data).hexdigest()

def get_title_with_retry(url: str, max_retries: int = 3) -> None | tuple[None, None] | tuple[str, BeautifulSoup]:
    """获取文章标题，带有重试机制"""
    retries = 0
    while retries < max_retries:
        try:
            # 设置请求头模拟浏览器访问
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
            }

            response = requests.get(url, headers=headers, timeout=10)
            response.raise_for_status()

            soup = BeautifulSoup(response.text, 'lxml')
            title_element = soup.find('h1', id="activity-name")
            if not title_element:
                # 尝试其他可能的标题位置
                title_element = soup.find('h2', class_="rich_media_title") or \
                                soup.find('h1', class_="article-title")

            if title_element:
                title = title_element.text.strip()
                if title:  # 确保标题不为空
                    return title, soup

            raise AttributeError("Title element not found")

        except Exception as e:
            retries += 1
            error_msg = str(e)
            if retries < max_retries:
                print(f"Retrying ({retries}/{max_retries}) for URL {url}: Error - {error_msg}")
                time.sleep(2 ** retries)  # 指数退避
            else:
                print(
                    f"Failed to retrieve title for URL {url} after {max_retries} retries. Last error: {error_msg}")
                return None, None

def process_url(url_data: Dict, output_base: str) -> bool:
    """处理单个URL数据"""
    url = url_data['url']
    account_name = url_data.get('account', 'unknown_account') or 'unknown_account'

    title, soup = get_title_with_retry(url)
    if not title or not soup:
        return False
    # 使用CSV中的日期或从网页提取
    create_time = url_data.get('date', '')
    if not create_time:
        match = re.search(r'createTime\s+=\s+\'(.*)\'', str(soup))
        if match:
            create_time = match.group(1)

    # 解析时间字符串为 datetime 对象
    time_obj = datetime.strptime(create_time, "%Y-%m-%d %H:%M")
    # 获取年份和月份
    year = time_obj.year
    month = time_obj.month
    # 输出结果
    print(f"年份: {year}")
    print(f"月份: {month}")
    # 创建公众号专属目录
    account_dir = os.path.join(output_base, account_name, str(year), str(month))
    os.makedirs(account_dir, exist_ok=True)

    create_time = f"publish_time: {create_time}" if create_time else "publish_time: unknown"

    content_soup = soup.find('div', {'class': 'rich_media_content'})
    if not content_soup:
        return False
    markdown, clean_title = convert_to_markdown(
        url, title, create_time, content_soup, account_dir)

    # 使用CSV中的标题或网页标题
    filename = url_data.get('title', clean_title)
    filename = re.sub(r'[\\/*?:"<>|]', '', filename)
    filepath = os.path.join(account_dir, f"{filename}.md")

    with open(filepath, 'w', encoding='utf-8') as f:
        f.write(markdown)

    print(f'Processed: {account_name} - {filename}.md')
    return True

def convert_to_markdown(url: str, title: str, create_time: str,
                        content_soup: BeautifulSoup, account_dir: str) -> tuple:
    """将HTML内容转换为Markdown格式"""
    download_images(content_soup, account_dir)
    markdown_content = markdownify.markdownify(str(content_soup))
    # 处理图片路径（确保使用相对路径）
    markdown_soup = BeautifulSoup(markdown_content, 'html.parser')
    for img in markdown_soup.find_all('img'):
        if img.get('data-src') and not img['data-src'].startswith(('http://', 'https://')):
            img['src'] = img['data-src']
    markdown_content = '\n'.join([line + '\n' for line in markdown_content.split('\n') if line.strip()])
    clean_title = remove_nonvisible_chars(title)

    markdown = f'# {clean_title}\n\n{create_time}\n\nurl: {url}\n\n{markdown_content}\n'

    markdown = re.sub('\xa0{1,}', '\n', markdown, flags=re.UNICODE)
    markdown = re.sub(r'\]\(http([^)]*)\)',
                      lambda x: '](http' + x.group(1).replace(' ', '%20') + ')',
                      markdown)
    return markdown, clean_title

def download_images(soup: BeautifulSoup, account_dir: str) -> None:
    """下载文章中的所有图片并保存到本地"""
    # 创建images目录（与markdown文件同级）
    image_folder = os.path.join(account_dir, 'images')
    os.makedirs(image_folder, exist_ok=True)
    for img in soup.find_all('img'):
        img_link = img.get('data-src') or img.get('src')
        if not img_link:
            continue
        img_link = img_link.replace(' ', '%20')

        if not img_link.startswith(('http://', 'https://')):
            img_link = 'https://mp.weixin.qq.com' + img_link
        try:
            with requests.get(img_link, stream=True) as response:
                response.raise_for_status()
                file_content = response.content

                img_hash = hash_byte_data(file_content)
                #if img_hash in filter_config.image_hashes:
                #    continue
                file_ext = (img.get('data-type') or 'jpg').split('?')[0]
                filename = f"{img_hash}.{file_ext}"
                filepath = os.path.join(image_folder, filename)

                if not os.path.exists(filepath):
                    with open(filepath, 'wb') as f:
                        f.write(file_content)
                # 更新图片属性指向本地文件（使用相对路径 ./images/）
                relative_path = f"./images/{filename}"
                img['data-src'] = relative_path
                img['src'] = relative_path
        except requests.exceptions.RequestException as e:
            print(f"图片下载失败，URL: {img_link}, 错误: {e}")


def remove_nonvisible_chars(text: str) -> str:
    return ''.join(c for c in text if (unicodedata.category(c) != 'Cn'
                                       and c not in (' ', '\n', '\r')))



def savetolist(resultPath, account, curl, title, result):
    wb = openpyxl.load_workbook(resultPath)
    ws = wb.active
    srow = [account, title, curl, result]
    ws.append(srow)
    wb.save(resultPath)


if __name__ == "__main__":

    output_dir = "D:\\WorkSpace\\Idea\\wechat"
    result_path = output_dir + "\\wx-result.xlsx"

    # 读取 Excel 文件
    file_path = output_dir + "\\wxlist.xlsx"

    wb_save = openpyxl.load_workbook(file_path)
    ws_save = wb_save.active

    wx_list = []
    # 读取每一行的值
    for row in ws_save.iter_rows(values_only=True):
        wx_list.append({
            'account': '禾木AI笔记-all',
            'title': row[0],
            'url': row[1]
            # 'date': row.get('日期', '').strip()
        })

    #print(wx_list)
    for item in wx_list:
        rs = 'success'
        try:
            time.sleep(random.randint(5, 10))
            process_url(item, output_dir)
        except Exception as e:
            # 捕获异常并处理（例如打印错误信息）
            print(f"发生错误: {e}")
            rs = str(e)
        # 文件路径
        # 写入 Excel
        savetolist(result_path, item['account'], item['url'], item['title'], rs)
