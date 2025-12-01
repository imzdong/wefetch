import requests
import time
import config
from openpyxl import Workbook
import random
url = config.articles_url
Cookie = config.cookie
headers = config.headers
token = config.token#公众号
fakeid = config.fakeid#公众号对应的id
type = '9'
data1 = {
    "token": token,
    "lang": "zh_CN",
    "f": "json",
    "ajax": "1",
    "action": "list_ex",
    "begin": "0",
    "count": "4",
    "query": "",
    "fakeid": fakeid,
    "type": type,
}
def getMoreInfo(link):
    # 获得mid,_biz,idx,sn 这几个在link中的信息
    mid = link.split("&")[1].split("=")[1]
    idx = link.split("&")[2].split("=")[1]
    sn = link.split("&")[3].split("=")[1]
    _biz = link.split("&")[0].split("_biz=")[1]
    pass_ticket = "fiddler中获取"#从fiddler中获取
    appmsg_token = "fiddler中获取"#从fiddler中获取
    url = "http://mp.weixin.qq.com/mp/getappmsgext"#获取详情页的网址
    phoneCookie = "在fiddler中获取"
    headers = {
        "Cookie": phoneCookie,
        "User-Agent": "fiddler中获取"
    }
    data = {
        "is_only_read": "1",
        "is_temp_url": "0",
        "appmsg_type": "9",
        'reward_uin_count': '0'
    }
    params = {
        "__biz": _biz,
        "mid": mid,
        "sn": sn,
        "idx": idx,
        "key": "fiddler中获取",
        "pass_ticket": pass_ticket,
        "appmsg_token": appmsg_token,
        "uin": "MTUyNzExNzYy",
        "wxtoken": "777",
    }
    requests.packages.urllib3.disable_warnings()
    content = requests.post(url, headers=headers, data=data, params=params).json()
    # print(content["appmsgstat"]["read_num"], content["appmsgstat"]["like_num"])
    try:
        readNum = content["appmsgstat"]["read_num"]
        print("阅读数:"+str(readNum))
    except:
        readNum = 0
    try:
        likeNum = content["appmsgstat"]["like_num"]
        print("喜爱数:"+str(likeNum))
    except:
        likeNum = 0
    try:
        old_like_num = content["appmsgstat"]["old_like_num"]
        print("在读数:"+str(old_like_num))
    except:
        old_like_num = 0
    time.sleep(3) # 歇3s，防止被封
    return readNum, likeNum,old_like_num
def getAllInfo(url):
    # 拿一页，存一页
    messageAllInfo = []
    # begin 从0开始
    for i in range(1):#设置爬虫页码
        begin = i * 4
        data1["begin"] = begin
        requests.packages.urllib3.disable_warnings()
        content_json = requests.get(url, headers=headers, params=data1, verify=False).json()
        time.sleep(random.randint(1, 10))
        if "app_msg_list" in content_json:
            for item in content_json["app_msg_list"]:
                spider_url = item['link']
                readNum, likeNum,old_like_num = getMoreInfo(spider_url)
                info = {
                    "title": item['title'],
                    "url": item['link'],
                    "readNum": readNum,
                    "likeNum": likeNum,
                    "old_like_num":old_like_num
                }
                messageAllInfo.append(info)
    return messageAllInfo
def main():
    f = Workbook()  # 创建一个workbook 设置编码
    sheet = f.active#创建sheet表单
    # 写入表头
    sheet.cell(row=1, column=1).value = 'title'  # 第一行第一列
    sheet.cell(row=1, column=2).value = 'url'
    sheet.cell(row=1, column=3).value = 'readNum(阅读数)'
    sheet.cell(row=1, column=4).value = 'likeNum(喜爱数)'
    sheet.cell(row=1, column=5).value = 'old_like_num(在看数)'
    messageAllInfo = getAllInfo(url)#获取信息
    print(messageAllInfo)
    print(len(messageAllInfo))#输出列表长度
    # 写内容
    for i in range(1, len(messageAllInfo)+1):
        sheet.cell(row=i + 1, column=1).value = messageAllInfo[i - 1]['title']
        sheet.cell(row=i + 1, column=2).value = messageAllInfo[i - 1]['url']
        sheet.cell(row=i + 1, column=3).value = messageAllInfo[i - 1]['readNum']
        sheet.cell(row=i + 1, column=4).value = messageAllInfo[i - 1]['likeNum']
        sheet.cell(row=i + 1, column=5).value = messageAllInfo[i - 1]['old_like_num']
    f.save(u'公众号.xls')  # 保存文件
if __name__ == '__main__':
    main()


'''
5.运行Fiddler，打开微信公众号客户端，找到已关注订阅列表，打开该公众号的历史文章页面，并随便点开一篇文章。
6.找到带有mp/getappmsgext?的请求包，在「Params」里找到「Key」、「Pass_ticket」、「appmasg_token」，在「Headers」里找到「Cookie」和「User-Agent」这五个数据也是我们所需要的。
'''